from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import subprocess
import sys
import xml.etree.ElementTree as ET
from collections import Counter, defaultdict
from copy import copy
from datetime import datetime, timedelta, timezone
from pathlib import Path
from urllib.parse import quote_plus
from urllib.request import Request, urlopen
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from openpyxl.styles import Alignment


PROJECT_ROOT = Path(__file__).resolve().parents[1]
WORKSPACE_ROOT = PROJECT_ROOT.parent
DEFAULT_TASK_ROOT = WORKSPACE_ROOT / "Sugar News"
PUBLIC_ROOT = PROJECT_ROOT / "public" / "sugar-news"
PUBLIC_DATA_ROOT = PUBLIC_ROOT / "data"
try:
    SHANGHAI = ZoneInfo("Asia/Shanghai")
except Exception:
    SHANGHAI = timezone(timedelta(hours=8), name="Asia/Shanghai")
GROUP_ORDER = {"巴西": 0, "印度": 1, "泰国": 2, "其他国家": 3}
IMPACT_PREFIXES = ("偏多糖价：", "偏空糖价：", "中性：", "影响有限：")
PLACEHOLDERS = (
    "暂无新闻",
    "暂无最新数据",
    "暂无最新对比数据",
    "暂无可比数据",
    "暂无最新",
    "暂未更新",
    "数据尚未公布",
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build Sugar News Excel and dashboard data.")
    parser.add_argument("--date", help="Target news date in YYYY-MM-DD. Defaults to Beijing yesterday.")
    parser.add_argument("--task-root", help="Existing Sugar News task root. Defaults to ../Sugar News.")
    parser.add_argument("--skip-if-success", action="store_true", help="Skip if public status already marks target date successful.")
    parser.add_argument("--offline-only", action="store_true", help="Do not attempt fallback online discovery; require verified JSON.")
    return parser.parse_args()


def beijing_now() -> datetime:
    fixed = os.getenv("SUGAR_NEWS_NOW")
    if fixed:
        return datetime.fromisoformat(fixed).astimezone(SHANGHAI)
    return datetime.now(SHANGHAI)


def target_date(value: str | None) -> str:
    if value:
        return datetime.strptime(value, "%Y-%m-%d").date().isoformat()
    return (beijing_now().date() - timedelta(days=1)).isoformat()


def task_root_from_args(value: str | None) -> Path:
    root = Path(value or os.getenv("SUGAR_NEWS_ROOT", str(DEFAULT_TASK_ROOT))).resolve()
    return root


def date_parts(date_text: str) -> tuple[str, str]:
    yyyy, mm, _ = date_text.split("-")
    return yyyy, mm


def ensure_task_dirs(task_root: Path, date_text: str) -> None:
    yyyy, mm = date_parts(date_text)
    for rel in (
        Path("data") / "verified_news" / yyyy / mm,
        Path("logs") / yyyy / mm,
        Path("reports") / yyyy / mm,
    ):
        (task_root / rel).mkdir(parents=True, exist_ok=True)


def verified_json_path(task_root: Path, date_text: str) -> Path:
    yyyy, mm = date_parts(date_text)
    return task_root / "data" / "verified_news" / yyyy / mm / f"sugar_news_{date_text}.json"


def search_log_path(task_root: Path, date_text: str) -> Path:
    yyyy, mm = date_parts(date_text)
    return task_root / "logs" / yyyy / mm / f"search_log_{date_text}.json"


def write_log_path(task_root: Path, date_text: str) -> Path:
    yyyy, mm = date_parts(date_text)
    return task_root / "logs" / yyyy / mm / f"write_log_{date_text}.json"


def excel_path(task_root: Path, date_text: str) -> Path:
    yyyy, mm = date_parts(date_text)
    return task_root / "reports" / yyyy / mm / f"Sugar News {date_text}.xlsx"


def public_report_path(date_text: str) -> Path:
    yyyy, mm = date_parts(date_text)
    return PUBLIC_DATA_ROOT / "reports" / yyyy / mm / f"{date_text}.json"


def public_index_path() -> Path:
    return PUBLIC_DATA_ROOT / "index.json"


def public_status_path() -> Path:
    return PUBLIC_DATA_ROOT / "status.json"


def success_exists(date_text: str) -> bool:
    path = public_status_path()
    if not path.exists():
        return False
    with path.open("r", encoding="utf-8") as f:
        status = json.load(f)
    return status.get("latestNewsDate") == date_text and status.get("lastRunStatus") == "success"


def google_news_rss_url(query: str) -> str:
    return "https://news.google.com/rss/search?q=" + quote_plus(query) + "&hl=en-US&gl=US&ceid=US:en"


def fetch_rss(query: str, timeout: int = 15) -> list[dict]:
    req = Request(google_news_rss_url(query), headers={"User-Agent": "Mozilla/5.0"})
    with urlopen(req, timeout=timeout) as resp:
        body = resp.read()
    root = ET.fromstring(body)
    items = []
    for node in root.findall("./channel/item"):
        title = node.findtext("title") or ""
        link = node.findtext("link") or ""
        published = node.findtext("pubDate") or ""
        desc = node.findtext("description") or ""
        items.append({"title": title, "link": link, "published": published, "description": desc})
    return items


def fallback_discovery(date_text: str, task_root: Path) -> None:
    """Record auditable search attempts.

    The cloud job needs a durable trail even when a fully verified newsroom-style
    dataset cannot be produced automatically. This fallback intentionally does
    not publish unverified RSS items as facts.
    """
    yyyy, mm = date_parts(date_text)
    dt = datetime.strptime(date_text, "%Y-%m-%d")
    buddhist_year = dt.year + 543
    readable = dt.strftime("%B %-d %Y") if os.name != "nt" else dt.strftime("%B %#d %Y")
    searches = [
        ("巴西", "en", f"Brazil sugar industry news {readable}"),
        ("巴西", "en", f"Brazil sugarcane ethanol export {readable}"),
        ("巴西", "pt-BR", f"Brasil açúcar etanol {dt.day} julho {dt.year}"),
        ("巴西", "pt-BR", f"Brasil setor sucroenergético {dt.day} de julho de {dt.year}"),
        ("印度", "en", f"India sugar industry news {readable}"),
        ("印度", "hi", f"भारत चीनी उद्योग {dt.day} जुलाई {dt.year}"),
        ("泰国", "en", f"Thailand sugar industry news {readable}"),
        ("泰国", "th", f"ประเทศไทย น้ำตาล อ้อย {dt.day} กรกฎาคม {buddhist_year}"),
        ("其他国家", "en", f"ICE sugar futures {readable}"),
    ]
    log = {
        "target_date": date_text,
        "run_date": beijing_now().date().isoformat(),
        "search_tool": "Google News RSS fallback via urllib",
        "note": "RSS search results are logged for audit. Items are not published unless a verified JSON is created.",
        "searches": [],
        "pipeline_counts": {
            "candidate_news_after_search": 0,
            "date_verified_or_continuing_impact": 0,
            "relevance_passed": 0,
            "deduped": 0,
            "passed_to_excel": 0,
        },
    }
    total = 0
    for country, language, query in searches:
        entry = {
            "country": country,
            "language": language,
            "keywords": query,
            "request_status": "pending",
            "returned_count": 0,
            "retained_count": 0,
            "filtered": [],
        }
        try:
            items = fetch_rss(query)
            entry["request_status"] = "executed"
            entry["returned_count"] = len(items)
            total += len(items)
            entry["sample_results"] = items[:5]
            entry["filtered"].append({"reason": "RSS result requires source-page date/body verification before publication."})
        except Exception as exc:
            entry["request_status"] = "failed"
            entry["error"] = str(exc)[:500]
        log["searches"].append(entry)
    log["pipeline_counts"]["candidate_news_after_search"] = total
    path = search_log_path(task_root, date_text)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(log, f, ensure_ascii=False, indent=2)


def load_verified_or_fail(task_root: Path, date_text: str, offline_only: bool) -> dict:
    path = verified_json_path(task_root, date_text)
    if not path.exists() and not offline_only:
        fallback_discovery(date_text, task_root)
    if not path.exists():
        raise FileNotFoundError(
            f"Missing verified Sugar News data: {path}. "
            "The job stopped before Excel/dashboard publication to avoid publishing blank or unverified content."
        )
    with path.open("r", encoding="utf-8") as f:
        data = json.load(f)
    if data.get("target_date") != date_text:
        raise ValueError(f"target_date mismatch in {path}")
    return data


def normalize_items(data: dict) -> list[dict]:
    items = data.get("items") or []
    seen = set()
    normalized = []
    for idx, item in enumerate(items, start=1):
        for field in ("country_group", "country", "news", "impact", "source_name", "source_url", "published_date_local"):
            if not item.get(field):
                raise ValueError(f"Verified item {idx} missing {field}")
        if not item["impact"].startswith(IMPACT_PREFIXES):
            raise ValueError(f"Verified item {idx} has invalid impact prefix")
        if any(text in item["news"] or text in item["impact"] for text in PLACEHOLDERS):
            raise ValueError(f"Verified item {idx} contains placeholder wording")
        if re.search(r"\bLMT\b|lmt", item["news"]):
            raise ValueError(f"Verified item {idx} contains raw LMT/lmt unit")
        if "来源：" not in item["news"] or item["source_url"] not in item["news"]:
            raise ValueError(f"Verified item {idx} missing B-column source link")
        if item["published_date_local"] != data["target_date"] and item.get("date_status") != "continuing_impact":
            raise ValueError(f"Verified item {idx} date is not target date or continuing impact")
        dedupe_key = item.get("dedupe_key") or re.sub(r"\s+", "", item["news"][:100])
        if dedupe_key in seen:
            raise ValueError(f"Duplicate verified news: {dedupe_key}")
        seen.add(dedupe_key)
        row = dict(item)
        row["_order"] = idx
        normalized.append(row)
    return sorted(normalized, key=lambda x: (GROUP_ORDER.get(x["country_group"], 3), -int(x.get("importance", 0)), x["_order"]))


def copy_row_style(source_ws, source_row: int, target_ws, target_row: int) -> None:
    for col in range(1, 4):
        source = source_ws.cell(source_row, col)
        target = target_ws.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        target.number_format = source.number_format
        target.protection = copy(source.protection)
        target.alignment = copy(source.alignment)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.font = copy(source.font)


def write_excel(task_root: Path, date_text: str, items: list[dict]) -> Path:
    template = task_root / "templates" / "新闻格式.xlsx"
    if not template.exists():
        raise FileNotFoundError(f"Missing template: {template}")
    out = excel_path(task_root, date_text)
    out.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template, out)

    wb = load_workbook(out)
    ws = wb.active
    if [ws.cell(1, c).value for c in range(1, 4)] != ["国家", "新闻", "影响"]:
        raise ValueError("Excel template headers must be 国家/新闻/影响")
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    template_wb = load_workbook(template)
    template_ws = template_wb.active
    source_row = 2 if template_ws.max_row >= 2 else 1
    for row, item in enumerate(items, start=2):
        copy_row_style(template_ws, source_row, ws, row)
        ws.cell(row, 1).value = item["country"]
        ws.cell(row, 2).value = item["news"]
        ws.cell(row, 3).value = item["impact"]
        ws.cell(row, 1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row, 2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.cell(row, 3).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = max(72, min(180, 24 + 0.55 * max(len(item["news"]), len(item["impact"]))))
    for col in ("B", "C"):
        ws.column_dimensions[col].width = max(ws.column_dimensions[col].width or 0, 55)
    wb.save(out)
    return out


def read_excel_rows(path: Path) -> list[dict]:
    wb = load_workbook(path)
    ws = wb.active
    rows = []
    for row in range(2, ws.max_row + 1):
        country = ws.cell(row, 1).value
        news = ws.cell(row, 2).value
        impact = ws.cell(row, 3).value
        if country or news or impact:
            rows.append({"row": row, "country": country, "news": news, "impact": impact})
    return rows


def split_impact(value: str) -> tuple[str, str]:
    for prefix in IMPACT_PREFIXES:
        if value.startswith(prefix):
            return prefix[:-1], value[len(prefix):]
    raise ValueError(f"Invalid impact value: {value}")


def build_dashboard_payload(date_text: str, items: list[dict], excel_file: Path) -> dict:
    grouped: dict[str, list[dict]] = defaultdict(list)
    country_order: list[tuple[int, str]] = []
    for item in items:
        impact_type, impact_text = split_impact(item["impact"])
        grouped[item["country"]].append({
            "news": re.sub(r"\s*来源：.*$", "", item["news"]).strip(),
            "impactType": impact_type,
            "impact": impact_text.strip(),
            "sourceName": item["source_name"],
            "sourceUrl": item["source_url"],
            "publishedDateLocal": item["published_date_local"],
            "eventDate": item.get("event_date"),
        })
        country_order.append((GROUP_ORDER.get(item["country_group"], 3), item["country"]))

    countries = []
    seen = set()
    for _, country in sorted(country_order, key=lambda pair: (pair[0], country_order.index(pair))):
        if country in seen:
            continue
        seen.add(country)
        if grouped[country]:
            countries.append({"country": country, "items": grouped[country]})

    return {
        "newsDate": date_text,
        "updatedAt": beijing_now().isoformat(timespec="seconds"),
        "timezone": "Asia/Shanghai",
        "excelFile": str(excel_file.relative_to(WORKSPACE_ROOT)).replace("\\", "/"),
        "countries": countries,
    }


def write_dashboard_data(date_text: str, payload: dict) -> tuple[Path, Path]:
    report_path = public_report_path(date_text)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    with report_path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    reports = []
    reports_root = PUBLIC_DATA_ROOT / "reports"
    if reports_root.exists():
        for path in reports_root.rglob("*.json"):
            with path.open("r", encoding="utf-8") as f:
                entry = json.load(f)
            reports.append({
                "newsDate": entry["newsDate"],
                "updatedAt": entry.get("updatedAt"),
                "path": "/" + str(path.relative_to(PROJECT_ROOT)).replace("\\", "/"),
                "count": sum(len(c.get("items", [])) for c in entry.get("countries", [])),
            })
    reports.sort(key=lambda x: x["newsDate"], reverse=True)
    index = {
        "latestNewsDate": reports[0]["newsDate"] if reports else None,
        "updatedAt": beijing_now().isoformat(timespec="seconds"),
        "reports": reports,
    }
    index_path = public_index_path()
    index_path.parent.mkdir(parents=True, exist_ok=True)
    with index_path.open("w", encoding="utf-8") as f:
        json.dump(index, f, ensure_ascii=False, indent=2)
    return report_path, index_path


def validate_all(date_text: str, items: list[dict], excel_file: Path, report_path: Path, index_path: Path) -> dict:
    excel_rows = read_excel_rows(excel_file)
    expected_pairs = {(item["country"], item["news"]) for item in items}
    actual_pairs = {(row["country"], row["news"]) for row in excel_rows}
    if expected_pairs != actual_pairs:
        missing = expected_pairs - actual_pairs
        extra = actual_pairs - expected_pairs
        raise ValueError(f"Excel mismatch: missing={missing}; extra={extra}")

    with report_path.open("r", encoding="utf-8") as f:
        report = json.load(f)
    with index_path.open("r", encoding="utf-8") as f:
        index = json.load(f)
    dashboard_count = sum(len(c.get("items", [])) for c in report.get("countries", []))
    if dashboard_count != len(items):
        raise ValueError(f"Dashboard count mismatch: {dashboard_count} != {len(items)}")
    if report.get("newsDate") != date_text:
        raise ValueError("Dashboard report date mismatch")
    if index.get("latestNewsDate") < date_text:
        raise ValueError("Dashboard index latest date is older than target date")
    if any(not c.get("items") for c in report.get("countries", [])):
        raise ValueError("Dashboard contains empty country section")

    group_positions = []
    for row in excel_rows:
        if row["country"] == "巴西":
            group_positions.append(0)
        elif row["country"] == "印度":
            group_positions.append(1)
        elif row["country"] == "泰国":
            group_positions.append(2)
        else:
            group_positions.append(3)
    checks = {
        "verified_count": len(items),
        "excel_count": len(excel_rows),
        "dashboard_count": dashboard_count,
        "excel_matches_verified": True,
        "dashboard_matches_verified": True,
        "country_order_ok": group_positions == sorted(group_positions),
        "no_empty_country_sections": True,
        "counts_by_country": dict(Counter(item["country"] for item in items)),
    }
    return checks


def write_status(date_text: str, status: str, details: dict, error: str | None = None) -> None:
    path = public_status_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "latestNewsDate": date_text if status == "success" else None,
        "lastRunAt": beijing_now().isoformat(timespec="seconds"),
        "lastRunStatus": status,
        "timezone": "Asia/Shanghai",
        "details": details,
    }
    if error:
        payload["error"] = error[:1000]
    if path.exists() and status != "success":
        with path.open("r", encoding="utf-8") as f:
            old = json.load(f)
        payload["latestNewsDate"] = old.get("latestNewsDate")
    with path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def write_task_log(task_root: Path, date_text: str, payload: dict) -> None:
    path = write_log_path(task_root, date_text)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def main() -> int:
    args = parse_args()
    date_text = target_date(args.date)
    task_root = task_root_from_args(args.task_root)
    ensure_task_dirs(task_root, date_text)

    if args.skip_if_success and success_exists(date_text):
        print(json.dumps({"status": "skipped", "reason": "already_success", "newsDate": date_text}, ensure_ascii=False))
        return 0

    try:
        data = load_verified_or_fail(task_root, date_text, offline_only=args.offline_only)
        items = normalize_items(data)
        excel_file = write_excel(task_root, date_text, items)
        payload = build_dashboard_payload(date_text, items, excel_file)
        report_path, index_path = write_dashboard_data(date_text, payload)
        checks = validate_all(date_text, items, excel_file, report_path, index_path)
        log_payload = {
            "target_date": date_text,
            "generated_at": beijing_now().isoformat(timespec="seconds"),
            "status": "success",
            "verified_news_file": str(verified_json_path(task_root, date_text)),
            "excel_file": str(excel_file),
            "dashboard_report": str(report_path),
            "dashboard_index": str(index_path),
            "checks": checks,
        }
        write_task_log(task_root, date_text, log_payload)
        write_status(date_text, "success", checks)
        print(json.dumps(log_payload, ensure_ascii=False, indent=2))
        return 0
    except Exception as exc:
        error = str(exc)
        details = {"target_date": date_text, "task_root": str(task_root)}
        write_task_log(task_root, date_text, {
            "target_date": date_text,
            "generated_at": beijing_now().isoformat(timespec="seconds"),
            "status": "failed",
            "error": error,
        })
        write_status(date_text, "failed", details, error=error)
        print(json.dumps({"status": "failed", "newsDate": date_text, "error": error}, ensure_ascii=False, indent=2), file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
